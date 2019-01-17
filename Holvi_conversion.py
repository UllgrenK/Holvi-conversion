#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 19:26:05 2019

@author: kristiina
"""

import openpyxl
import os

os.chdir('/Users/Kristiina/KP')
wb = openpyxl.load_workbook('2018_kp.xlsx')
Alkup = wb['Alkup']
Holvi = wb.create_sheet("Holvi")
max_row = Alkup.max_row
max_column = Alkup.max_column
row_counter = 0

def getDebet(Sum, row):
    if Sum < 0:
        return (Alkup.cell(row,column=4).value)[0:4]
    else: 
        return '1710'
    
def getKredit(Sum, row):
    if Sum < 0:
        return '1710'
    else:
        return (Alkup.cell(row,column=4).value)[0:4]

List = []
Tositenro = 0
for i in range(8,max_row+1):
    Tositenro = Tositenro +1
    Pvm = (Alkup.cell(row=i,column=1).value)[0:-5]
    Selite1 = (Alkup.cell(row=i,column=6).value)
    Selite2 = 'None'
    Summa = (Alkup.cell(row=i,column=12).value)
    Debet = getDebet(Summa, i)
    Kredit = getKredit(Summa, i)
    Summa = abs(Summa)
    List = [Tositenro, Pvm, Selite1, Selite2, Debet, Kredit, Summa]
    print(List)
    Holvi.append(List)
    
    
    
    #poimi tarpeelliset solut, muokkaa, laittaa listaan
    #kirjoitta tuloksen Holvi-sheetille rivi kerrallaan
    
    
#talleta työkirja
wb.save('2018_kp.xlsx')
